import streamlit as st
import datetime as dt
from tempfile import NamedTemporaryFile
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from operator import itemgetter

def main():
    st.set_page_config(page_title='SCC-OPS', layout='wide', initial_sidebar_state='expanded')
    oncall_schedules = st.sidebar.checkbox('On-Call Schedules', value=False)
    if oncall_schedules:
        get_oncall_schedules()

def get_oncall_schedules():
    def download_excel_file(wb, filename):
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            data = BytesIO(tmp.read())
        st.download_button('download schedule', data = data, mime='xlsx', file_name=filename)

    def get_names(name):
        if name == 'OPEN':
            return('Open', ' ')
        if ',' in name:
            new_name = name.split(', ')
            fname = new_name[1]
            lname = new_name[0]
            return (fname, lname)
        return ('unknown', 'unknown')

    def get_full_name(name):
        fname, lname = get_names(name)
        full_name = fname + ' ' + lname
        return full_name

    def get_times(time):
        new_time = time.split('-')
        start_time = new_time[0]
        start_time_standard = dt.datetime.strptime(start_time, '%H:%M').strftime('%I:%M %p')
        end_time = new_time[1]
        end_time_standard = dt.datetime.strptime(end_time, '%H:%M').strftime('%I:%M %p')
        return (start_time, end_time, start_time_standard, end_time_standard)

    def format_date(date):
        weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday', 'Monday']
        months = [' ', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

        split_date = date.split('-')
        year = split_date[0]
        month = months[int(split_date[1])]
        day = str(int(split_date[2]))

        converted_date = dt.datetime.strptime(date, '%Y-%m-%d')
        weekday_num = converted_date.weekday()
        weekday = weekdays[weekday_num]
        new_date = weekday + ', ' + month + ' ' + day + ', ' + year
        return new_date

    def order_by_time(dates_dict):
        new_dates_dict = {}
        for d, v in dates_dict.items():
            new_v = sorted(v, key=itemgetter(1))
            new_dates_dict[d] = new_v
        # st.write(dates_dict)
        # st.write(new_dates_dict)
        return new_dates_dict

    def order_by_time_2(dates_dict):
        new_dates_dict = {}
        for d, v in dates_dict.items():
            new_v = sorted(v, key=itemgetter(2))
            new_dates_dict[d] = new_v
        # st.write(dates_dict)
        # st.write(new_dates_dict)
        return new_dates_dict

    def get_data_from_file(fl):
        data = []
        wb = load_workbook(fl)
        ws = wb.active
        dates = []
        dates_dict = {}
        date = ''
        row = 2
        while date is not None:
            row_data = []        
            date_cell = 'A' + str(row)
            date = ws[date_cell].value
            if date is None:
                break
            date = format_date(date)
            if date not in dates:
                dates.append(date)
                dates_dict[date] = []
            name_cell = 'D' + str(row)
            name = ws[name_cell].value
            # fname, lname = get_names(name)
            # row_data.append(fname)
            # row_data.append(lname)
            fullname = get_full_name(name)
            row_data.append(fullname)
            time_cell = 'E' + str(row)
            time = ws[time_cell].value
            start_time, end_time, start_time_standard, end_time_standard = get_times(time)
            row_data.append(start_time)
            row_data.append(end_time)
            row_data.append(start_time_standard)
            row_data.append(end_time_standard)
            description_cell = 'G' + str(row)
            description = ws[description_cell].value
            row_data.append(description)
            hours_cell = 'H' + str(row)
            hours = ' '#ws[hours_cell].value
            row_data.append(hours)
            notes_cell = 'I' + str(row)
            notes = ws[notes_cell].value
            row_data.append(notes)

            dates_dict[date].append(row_data)
            row += 1
        new_dates_dict = order_by_time(dates_dict)
        return new_dates_dict

    def apply_style_merged_cell(ws, row, cols, font, bold, color):
        for col in cols:
            cell = col + str(row)
            ws[cell].font = Font(size=font, bold=bold)
            ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            ws[cell].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_style_merged_cell_2(ws, row, cols, font, bold, color):
        for col in cols:
            cell = col + str(row)
            ws[cell].font = Font(size=font, bold=bold)
            ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[cell].alignment = Alignment(horizontal='left', vertical='center')
            ws[cell].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    def apply_style_cell(ws, row, cols, font, bold, color):
        for col in cols:
            cell = col + str(row)
            ws[cell].font = Font(size=font, bold=bold)
            ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            ws[cell].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_style_shift_row(ws, row):
        cell = 'A' + str(row)
        ws[cell].alignment = Alignment(horizontal='left')

    def apply_strike(ws, row):
        for col in 'ABCDE':
            cell = col + str(row)
            ws[cell].font = Font(strike=True, color='00FF0000', size=13)

    def apply_strike_2(ws, row):
        for col in 'ABCDEFGHIJ':
            cell = col + str(row)
            ws[cell].font = Font(strike=True, color='00FF0000', size=13)

    def apply_style_shift_row_aata(ws, row):
        cell = 'A' + str(row)
        ws[cell].alignment = Alignment(horizontal='left')
        cell = 'F' + str(row)
        ws[cell].alignment = Alignment(horizontal='left')
        cell = 'D' + str(row)
        ws[cell].alignment = Alignment(horizontal='left')
        cell = 'E' + str(row)
        ws[cell].alignment = Alignment(horizontal='right')

    def apply_style_shift_row_aata_2(ws, row):
        cell = 'A' + str(row)
        ws[cell].alignment = Alignment(horizontal='left')
        cell = 'C' + str(row)
        ws[cell].alignment = Alignment(horizontal='center')
        cell = 'D' + str(row)
        ws[cell].alignment = Alignment(horizontal='center')
        cell = 'J' + str(row)
        ws[cell].alignment = Alignment(horizontal='right')

    def create_xl_file(data):
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16

        row = 1
        #ws = wb.create_sheet('Schedule')
        cols = ['A', 'B', 'C', 'D', 'E']
        ws.merge_cells(f'A{row}:E{row}')
        cell = f'A{row}'
        ws[cell] = 'Operations Department On-Call Staff Schedule'
        apply_style_merged_cell(ws, row, cols, 20, True, 'FF76B5C5')
        row += 1

        ws.append(['Name','Start Time','End Time', 'Description', 'Notes'])
        apply_style_cell(ws, row, cols, 18, True, 'FF76B5C5')
        row += 1

        for d,v in data.items():
            ws.merge_cells(f'A{row}:E{row}')
            cell = f'A{row}'
            ws[cell] = d
            apply_style_merged_cell(ws, row, cols, 16, True, 'FFABDBE3')

            for i in v:
                row +=1
                new_row = [i[0], i[3], i[4], i[5], i[6]]
                ws.append(new_row)
                apply_style_cell(ws, row, cols, 13, False, 'FFFFFFFF')
                if 'Open' in i[0]:
                    apply_style_cell(ws, row, cols, 13, False, 'FFFFFF00')
                if 'OUT' in i[7]:
                    apply_strike(ws,row)                
                apply_style_shift_row(ws, row)
            row +=1            

        return wb

    def create_report(fl):
        data = get_data_from_file(fl)
        dates = []
        for d in data:
            a = d.split(', ')[1].split(' ')
            b = a[0][:3]+a[1]
            dates.append(b)
        file_name = 'setup-'+dates[0]+'-'+dates[-1]+'.xlsx'
        st.write(f'Your file is ready.') 
        st.write(f'File name is {file_name}')
        wb = create_xl_file(data)
        download_excel_file(wb, file_name)

    def get_data_from_file_aata(fl):
        data = []
        wb = load_workbook(fl)
        ws = wb.active
        dates = []
        dates_dict = {}
        date = ''
        row = 2
        while date is not None:
            row_data = []        
            date_cell = 'A' + str(row)
            date = ws[date_cell].value
            if date is None:
                break
            date = format_date(date)
            if date not in dates:
                dates.append(date)
                dates_dict[date] = []
            name_cell = 'D' + str(row)
            name = ws[name_cell].value
            # fname, lname = get_names(name)
            # row_data.append(fname)
            # row_data.append(lname)
            fullname = get_full_name(name)
            row_data.append(fullname)
            time_cell = 'E' + str(row)
            time = ws[time_cell].value
            start_time, end_time, start_time_standard, end_time_standard = get_times(time)
            row_data.append(start_time)
            row_data.append(end_time)
            row_data.append(start_time_standard)
            row_data.append(end_time_standard)
            description_cell = 'G' + str(row)
            description = ws[description_cell].value
            row_data.append(description)
            hours_cell = 'H' + str(row)
            hours = ws[hours_cell].value
            row_data.append(hours)
            notes_cell = 'I' + str(row)
            notes = ws[notes_cell].value
            row_data.append(notes)

            dates_dict[date].append(row_data)
            row += 1
        new_dates_dict = order_by_time(dates_dict)
        return new_dates_dict

    def create_xl_file_aata(data):
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 50

        row = 1
        #ws = wb.create_sheet('Schedule')
        cols = ['A', 'B', 'C', 'D', 'E', 'F']
        ws.merge_cells(f'A{row}:F{row}')
        cell = f'A{row}'
        ws[cell] = 'Event Services Department On-Call Staff Schedule'
        apply_style_merged_cell(ws, row, cols, 20, True, 'FF76B5C5')
        row += 1

        ws.append(['Name','Start Time','End Time', 'Description', 'Hours', 'Notes'])
        apply_style_cell(ws, row, cols, 18, True, 'FF76B5C5')
        row += 1

        for d,v in data.items():
            ws.merge_cells(f'A{row}:F{row}')
            cell = f'A{row}'
            ws[cell] = d
            apply_style_merged_cell(ws, row, cols, 16, True, 'FFABDBE3')

            for i in v:
                row +=1
                new_row = [i[0], i[3], i[4], i[5], i[6], i[7]]
                ws.append(new_row)
                apply_style_cell(ws, row, cols, 13, False, 'FFFFFFFF')
                if 'Open' in i[0]:
                    apply_style_cell(ws, row, cols, 13, False, 'FFFFFF00')
                apply_style_shift_row_aata(ws, row)
            row +=1            

        return wb

    def create_report_aata(fl):
        data = get_data_from_file_aata(fl)
        dates = []
        for d in data:
            a = d.split(', ')[1].split(' ')
            b = a[0][:3]+a[1]
            dates.append(b)
        file_name = 'AATA-'+dates[0]+'-'+dates[-1]+'.xlsx'
        st.write(f'Your file is ready.') 
        st.write(f'File name is {file_name}')
        wb = create_xl_file_aata(data)
        download_excel_file(wb, file_name)

#BEGINNING OF AATA VERSION 2

    def get_data_from_file_aata_2(fl):
        data = []
        wb = load_workbook(fl)
        ws = wb.active
        dates = []
        dates_dict = {}
        date = ''
        row = 2
        while date is not None:
            row_data = []        
            date_cell = 'A' + str(row)
            date = ws[date_cell].value
            if date is None:
                break
            date = format_date(date)
            if date not in dates:
                dates.append(date)
                dates_dict[date] = []
            name_cell = 'D' + str(row)
            name = ws[name_cell].value
            # fname, lname = get_names(name)
            # row_data.append(fname)
            # row_data.append(lname)
            fullname = get_full_name(name)
            row_data.append(fullname)

            building_cell = 'E' + str(row)
            building = ws[building_cell].value
            row_data.append(building)

            time_cell = 'F' + str(row)
            time = ws[time_cell].value
            start_time, end_time, start_time_standard, end_time_standard = get_times(time)
            row_data.append(start_time)
            row_data.append(end_time)
            row_data.append(start_time_standard)
            row_data.append(end_time_standard)
            description_cell = 'K' + str(row)
            description = ws[description_cell].value
            row_data.append(description)
            hours_cell = 'I' + str(row)
            hours = ws[hours_cell].value
            row_data.append(hours)
            notes_cell = 'J' + str(row)
            notes = ws[notes_cell].value
            row_data.append(notes)

            dates_dict[date].append(row_data)
            row += 1
        new_dates_dict = order_by_time_2(dates_dict)
        return new_dates_dict

    def create_xl_file_aata_2(data):
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 10

        row = 1
        #ws = wb.create_sheet('Schedule')
        cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        ws.merge_cells(f'A{row}:J{row}')
        cell = f'A{row}'
        ws[cell] = 'Event Services Department On-Call Staff Schedule'
        apply_style_merged_cell(ws, row, cols, 18, True, 'FF76B5C5')
        row += 1

        ws.append(['Name','Building','Description','Notes','Start Time','End Time', 'Break', 'Meal', 'Break','Hours'])
        apply_style_cell(ws, row, cols, 18, True, 'FF76B5C5')
        row += 1

        for d,v in data.items():
            ws.merge_cells(f'A{row}:J{row}')
            cell = f'A{row}'
            ws[cell] = d
            apply_style_merged_cell_2(ws, row, cols, 16, True, 'FFABDBE3')

            for i in v:
                row +=1
                new_row = [i[0], i[1], i[6], i[8], i[4], i[5], ' ', ' ', ' ', i[7]]
                ws.append(new_row)
                apply_style_cell(ws, row, cols, 14, False, 'FFFFFFFF')
                if 'Open' in i[0]:
                    apply_style_cell(ws, row, cols, 14, False, 'FFFFFF00')
                if 'OUT' in i[8]:
                    apply_strike_2(ws,row)  
                apply_style_shift_row_aata_2(ws, row)
            row +=1            

        return wb

    def create_report_aata_2(fl):
        data = get_data_from_file_aata_2(fl)
        dates = []
        for d in data:
            a = d.split(', ')[1].split(' ')
            b = a[0][:3]+a[1]
            dates.append(b)
        file_name = 'AATA-'+dates[0]+'-'+dates[-1]+'.xlsx'
        st.write(f'Your file is ready.') 
        st.write(f'File name is {file_name}')
        wb = create_xl_file_aata_2(data)
        download_excel_file(wb, file_name)


#END OF AATA VERSION 2

#BEGINNING OF SETUP VERSION 2

    def get_data_from_file_2(fl):
        data = []
        wb = load_workbook(fl)
        ws = wb.active
        dates = []
        dates_dict = {}
        date = ''
        row = 2
        while date is not None:
            row_data = []        
            date_cell = 'A' + str(row)
            date = ws[date_cell].value
            if date is None:
                break
            date = format_date(date)
            if date not in dates:
                dates.append(date)
                dates_dict[date] = []
            name_cell = 'D' + str(row)
            name = ws[name_cell].value
            # fname, lname = get_names(name)
            # row_data.append(fname)
            # row_data.append(lname)
            fullname = get_full_name(name)
            row_data.append(fullname)

            building_cell = 'E' + str(row)
            building = ws[building_cell].value
            row_data.append(building)

            time_cell = 'F' + str(row)
            time = ws[time_cell].value
            start_time, end_time, start_time_standard, end_time_standard = get_times(time)
            row_data.append(start_time)
            row_data.append(end_time)
            row_data.append(start_time_standard)
            row_data.append(end_time_standard)
            description_cell = 'H' + str(row)
            description = ws[description_cell].value
            row_data.append(description)
            hours_cell = 'I' + str(row)
            hours = ws[hours_cell].value
            row_data.append(hours)
            notes_cell = 'J' + str(row)
            notes = ws[notes_cell].value
            row_data.append(notes)

            dates_dict[date].append(row_data)
            row += 1
        new_dates_dict = order_by_time_2(dates_dict)
        return new_dates_dict

    def create_xl_file_2(data):
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        row = 1
        #ws = wb.create_sheet('Schedule')
        cols = ['A', 'B', 'C', 'D', 'E', 'F']
        ws.merge_cells(f'A{row}:F{row}')
        cell = f'A{row}'
        ws[cell] = 'Operations Department On-Call Staff Schedule'
        apply_style_merged_cell(ws, row, cols, 20, True, 'FF76B5C5')
        row += 1

        ws.append(['Name','Start Time','End Time', 'Description', 'Building', 'Notes'])
        apply_style_cell(ws, row, cols, 18, True, 'FF76B5C5')
        row += 1

        for d,v in data.items():
            ws.merge_cells(f'A{row}:F{row}')
            cell = f'A{row}'
            ws[cell] = d
            apply_style_merged_cell(ws, row, cols, 16, True, 'FFABDBE3')

            for i in v:
                row +=1
                new_row = [i[0], i[4], i[5], i[6], i[1], ' ']
                ws.append(new_row)
                apply_style_cell(ws, row, cols, 13, False, 'FFFFFFFF')
                if 'Open' in i[0]:
                    apply_style_cell(ws, row, cols, 13, False, 'FFFFFF00')
                if 'OUT' in i[8]:
                    apply_strike(ws,row)                
                apply_style_shift_row(ws, row)
            row +=1            

        return wb

    def create_report_2(fl):
        data = get_data_from_file_2(fl)
        dates = []
        for d in data:
            a = d.split(', ')[1].split(' ')
            b = a[0][:3]+a[1]
            dates.append(b)
        file_name = 'setup-'+dates[0]+'-'+dates[-1]+'.xlsx'
        st.write(f'Your file is ready.') 
        st.write(f'File name is {file_name}')
        wb = create_xl_file_2(data)
        download_excel_file(wb, file_name)

#END OF SETUP VERSION 2


    fl = st.sidebar.file_uploader('Upload Excel File:', accept_multiple_files=False)
    if st.sidebar.button('Create Setup Schedule') and fl is not None:
        create_report(fl)
    if st.sidebar.button('Create AA/TA Schedule') and fl is not None:
        create_report_aata(fl)
    if st.sidebar.button('Create Setup Schedule - NEW') and fl is not None:
        create_report_2(fl)
    if st.sidebar.button('Create AA/TA Schedule - NEW') and fl is not None:
        create_report_aata_2(fl)
              
    st.sidebar.markdown('<hr>', unsafe_allow_html=True)

if __name__ == '__main__':
    main()
